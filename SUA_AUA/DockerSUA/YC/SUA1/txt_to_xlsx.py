# -*- coding: UTF-8 -*-
# coding: UTF-8 

import time
from datetime import datetime
import sys, traceback
import threading
import ast
import time, requests, random, uuid
import json, atexit, sys, requests, os, signal, sqlite3
import ntplib
from openpyxl import Workbook, load_workbook

date = "2020-04-23"
sh_name = date

if os.path.isfile('./data_device.xlsx'):
    txt_filename = date+".txt"
    f=open(txt_filename,'r+') #打開Txt文件 
    # read file
    book = load_workbook(filename = './data_device.xlsx')
    sheet = book.create_sheet(sh_name) 
    if "Sheet" in book.get_sheet_names():
        ws = book["Sheet"]
        book.remove(ws)
    sheet.append(["tx_time(UNIX)", "root_delay(in microseconds)", "hour", "minute", "second", "microsecond"])
    line=f.readline(); #讀取第一行文本 
    while line: 
        datalist=[] 
        datalist=line.split(',') #將字符串以「，」為標識分片 
        for i in range(0, len(datalist)):  
            datalist[i] = datalist[i].strip('\n') # 去除換行符
        sheet.append(datalist) #寫入excel文檔中 
        line=f.readline() #讀取下一行 
    
else:    
    book = Workbook() # this action will create a empty default sheet "Sheet"
    txt_filename = date+".txt"
    f=open(txt_filename,'r+') #打開Txt文件 
    sheet = book.create_sheet(sh_name) 
    if "Sheet" in book.get_sheet_names():
        ws = book["Sheet"]
        book.remove(ws)
    sheet.append(["tx_time(UNIX)", "root_delay(in microseconds)", "hour", "minute", "second", "microsecond"])
    line=f.readline(); #讀取第一行文本 
    while line: 
        datalist=[] 
        datalist=line.split(',') #將字符串以「，」為標識分片 
        for i in range(0, len(datalist)):  
            datalist[i] = datalist[i].strip('\n') # 去除換行符
        sheet.append(datalist) #寫入excel文檔中 
        line=f.readline() #讀取下一行 

book.save("data_device.xlsx")
book.close()
